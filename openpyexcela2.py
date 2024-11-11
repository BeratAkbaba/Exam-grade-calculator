import math
import openpyxl
dosya = openpyxl.load_workbook("./excell.xlsx")
sayfa= dosya["ofis Yazılımları test verisi"]
sayfa2=dosya["harf aralığı"]
liste=[]
sayac_katilacaklar, toplam, sonuc = 0, 0, 0

y=lambda x: x if x != None else int(0)
q=lambda x: float(x)
g=lambda x: "Geçti" if x!="FF" else "Kaldı"

def hbp_hesapla(vizeyi,finali):
    return math.ceil(vizeyi*0.4+finali*0.6)
def gercek_hbsi(vizeyi,bütü,finali):
    return math.ceil(vizeyi*0.4+bütü*0.6) if sayfa.cell(i,10).value!=None else math.ceil(vizeyi*0.4+finali*0.6)

def bagila_katilacaklari_hesapla(hbp,final):
    if hbp>= 10 and final!=None:
        return "Katilacak"
    else:
        return "Katilmayacak"
def stsapma(hbp):
    global  sOrtalamasi, sonuc
    if sayfa.cell(i, 13).value == "Katilacak":
        sonuc=((hbp-sOrtalamasi)**2)+ sonuc
    return math.sqrt(sonuc/sayac_katilacaklar)
     
def t_skor_bulma(ghbp):
    global sOrtalamasi 
    if ghbp>=10:
        if y(satir[8])!=0 or  y(satir[7])!=0:
            z=(ghbp-sOrtalamasi) / stsapmasonucu
            t=round((z*10)+50,2)
            return t
        else:
            return 0
    else:
        return 0
def harf_araligi_bul(altsinir,üstsinir,i):
    if sOrtalamasi>=altsinir and sOrtalamasi<=üstsinir:
        return i
         
for i in range(2,sayfa.max_row+1):
    for j in range(2,18):
        liste.append(sayfa.cell(i,j).value)
        sayfa.cell(i,11).value = y(sayfa.cell(i,7).value)+y(sayfa.cell(i,8).value)
        sayfa.cell(i,12).value = hbp_hesapla(sayfa.cell(i,11).value, y(sayfa.cell(i,9).value))
        sayfa.cell(i,14).value= gercek_hbsi(sayfa.cell(i,11).value,sayfa.cell(i,10).value,y(sayfa.cell(i,9).value))
        sayfa.cell(row=i, column=11, value=sayfa.cell(i,14).value)
        sayfa.cell(i, 13).value= bagila_katilacaklari_hesapla(sayfa.cell(i, 12).value, sayfa.cell(i, 9).value)
        
    if sayfa.cell(i, 13).value == "Katilacak":
        sayac_katilacaklar+=1
        toplam+= sayfa.cell(i,12).value
sOrtalamasi=toplam/sayac_katilacaklar
for i in range(2,sayfa.max_row+1):
    stsapmasonucu=math.ceil(stsapma(sayfa.cell(i,12).value))

print(sayac_katilacaklar)
print(toplam)
print("sınıf ortalaması :",sOrtalamasi)
print("stsapma :",stsapmasonucu)

liste = [liste[i:i+16] for i in range(0, len(liste), 16)]
for satir in liste:
    satir[13]= t_skor_bulma(satir[12])
for i in range(2,sayfa.max_row+1):
    sayfa.cell(row=i, column=12, value=t_skor_bulma(sayfa.cell(i, 14).value))

for satir in liste:
    if sOrtalamasi>=70:
        for j in range(4,sayfa2.max_column+1):
            if satir[12]>=q(sayfa2.cell(9,j).value):
                satir[15]= sayfa2.cell(1,j).value
                break
    else:
        if sayac_katilacaklar<30:
            for j in range(4,sayfa2.max_column+1):
                if satir[12]>=q(sayfa2.cell(10,j).value):
                    satir[15]= sayfa2.cell(1,j).value
                    break
        else:
            for j in range(4,sayfa2.max_column+1):
                if satir[13]>=q(sayfa2.cell(2,j).value):
                    if satir[12]<40:
                        satir[15]="FF"
                    else:
                        satir[15]= sayfa2.cell(1,j).value
                        break


for satir in liste:
    satir[14]=g(satir[15])

for i in range(2,sayfa.max_row+1):
    sayfa.cell(row=i, column=13, value=satir[14])
    sayfa.cell(row=i, column=14, value=satir[15])


for satir in liste:  
    print(satir)

dosya.save("./excell-yeni.xlsx")

   
"""      
for ogrenci in liste:
    
    ogrenci[0] = f'{ogrenci[0]:<13}'
    ogrenci[1] = f'{ogrenci[1]:<17}'
    ogrenci[2] = f'{ogrenci[2]:<13}'
    ogrenci[4] = f'{ogrenci[4]:<14}'
    ogrenci[5] = f'{y(ogrenci[5]):<3}'
    ogrenci[6] = f'{y(ogrenci[6]):<3}'
    ogrenci[7] = f'{y(ogrenci[7]):<3}'
    ogrenci[8] = f'{y(ogrenci[8]):<3}'
    ogrenci[9] = f'{y(ogrenci[9]):<3}'
    ogrenci[10] = f'{y(ogrenci[10]):<3}'
    ogrenci[11] = f'{ogrenci[11]:<12}'
    ogrenci[12] = f'{y(ogrenci[12]):<3}'
    ogrenci[13] = f'{y(ogrenci[13]):<5}'

for ogrenci in liste:
    print(ogrenci)
""" 